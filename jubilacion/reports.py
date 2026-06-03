from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .calculator import CalculationResult, ScenarioResult, money
from .legal_basis import LEGAL_BASIS


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "logo-ronquillo.png"
FORMULA_PENSION_PATH = Path(__file__).resolve().parent.parent / "assets" / "formula_pension_mensual.png"
FORMULA_GLOBAL_PATH = Path(__file__).resolve().parent.parent / "assets" / "formula_fondo_global.png"
FOOTER_NAME = "Abg. Mgtr. Ing. Pablo Ronquillo"
FOOTER_BRAND = "AHIL Legal Tech"
FOOTER_PHONE = "0986658162"
FOOTER_PLACE = "Quito - Ecuador"


class ReportPDF(FPDF):
    def footer(self) -> None:
        self.set_y(-16)
        self.set_draw_color(31, 78, 121)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(2)
        self.set_font("Arial", "I", 7.5)
        self.set_text_color(80, 80, 80)
        footer_text = (
            f"{FOOTER_NAME} | {FOOTER_BRAND} | Celular: {FOOTER_PHONE} | "
            f"{FOOTER_PLACE}"
        )
        self.cell(0, 4, _latin(footer_text), ln=True, align="C")
        self.set_font("Arial", "", 7)
        self.cell(0, 4, _latin(f"Pagina {self.page_no()}"), align="C")
        self.set_text_color(0, 0, 0)


def build_excel_report(result: CalculationResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"
    _setup_sheet(ws)

    if LOGO_PATH.exists():
        try:
            logo = ExcelImage(str(LOGO_PATH))
            logo.width = 78
            logo.height = 78
            ws.add_image(logo, "A1")
            ws.row_dimensions[1].height = 60
        except Exception:
            pass

    ws["B1"] = "Informe de calculo de Jubilacion Patronal"
    ws["B1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["B1"].fill = TITLE_FILL
    ws.merge_cells("B1:F1")

    data_rows = [
        ("Trabajador", result.entrada.trabajador),
        ("Identificacion", result.entrada.identificacion),
        ("Empleador", result.entrada.empleador),
        ("Cargo", result.entrada.cargo),
        ("Fecha nacimiento", result.entrada.fecha_nacimiento.isoformat()),
        ("Fecha ingreso", result.entrada.fecha_ingreso.isoformat()),
        ("Fecha salida", result.entrada.fecha_salida.isoformat()),
        ("Edad al determinar renta", result.edad_renta),
        ("Tiempo de servicio", result.tiempo_servicio),
        ("Elegibilidad", result.elegibilidad),
        ("Promedio anual ultimos 5 anos", result.promedio_anual_ultimos_5),
        ("Promedio mensual ultimo anio", result.promedio_mens_ultimo_anio),
    ]
    row = 3
    for label, value in data_rows:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        row += 1

    row += 1
    ws.cell(row, 1, "Escenario recomendado").font = Font(bold=True)
    row += 1
    _write_scenario_summary(ws, row, result.recomendado)

    scenarios = wb.create_sheet("Escenarios")
    _setup_sheet(scenarios)
    _write_scenarios_sheet(scenarios, result)

    inputs = wb.create_sheet("Datos")
    _setup_sheet(inputs)
    _write_inputs_sheet(inputs, result)

    sources = wb.create_sheet("Fuentes")
    _setup_sheet(sources)
    _write_sources_sheet(sources, result)

    formulas = wb.create_sheet("Formulas")
    _setup_sheet(formulas)
    _write_formulas_sheet(formulas)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def build_pdf_report(result: CalculationResult) -> bytes:
    pdf = ReportPDF()
    pdf.set_auto_page_break(auto=True, margin=22)
    pdf.add_page()
    _pdf_header(pdf)

    _pdf_section(pdf, "Datos del trabajador")
    for label, value in [
        ("Trabajador", result.entrada.trabajador),
        ("Identificacion", result.entrada.identificacion),
        ("Empleador", result.entrada.empleador),
        ("Cargo", result.entrada.cargo),
        ("Periodo laboral", f"{result.entrada.fecha_ingreso} a {result.entrada.fecha_salida}"),
        ("Edad / servicio", f"{result.edad_renta} anos / {result.tiempo_servicio} anos"),
        ("Elegibilidad", result.elegibilidad),
    ]:
        _pdf_pair(pdf, label, str(value))

    _pdf_section(pdf, "Criterio jurisprudencial aplicado")
    _pdf_multiline(
        pdf,
        "Resolucion CNJ No. 16-2025: para el calculo de jubilacion patronal, "
        "el fondo de reserva integra obligatoriamente el haber individual. "
        "En aplicacion del principio de favorabilidad, el escenario recomendado "
        "rebaja del fondo de jubilacion solo los valores de fondos de reserva "
        "pagados, entregados o depositados por el empleador. Los aportes "
        "patronales no se rebajan en ese escenario y se muestran unicamente "
        "como escenario comparativo."
    )
    _pdf_multiline(
        pdf,
        "Resolucion CNJ No. 04-2026: para el fondo global, cuando existe acuerdo "
        "de las partes, se aplican los acuerdos ministeriales vigentes a la fecha "
        "de terminacion de la relacion laboral. Por eso se usa el coeficiente "
        "global C2 del anio de cese o el C2 manual cuando corresponda."
    )

    _pdf_section(pdf, "Base legal del Codigo del Trabajo")
    for title, detail in LEGAL_BASIS[:6]:
        _pdf_multiline(pdf, f"{title}: {detail}")

    _pdf_formula_section(pdf)

    _pdf_section(pdf, "Resumen recomendado")
    _pdf_scenario(pdf, result.recomendado, result)

    _pdf_section(pdf, "Escenario comparativo")
    _pdf_scenario(pdf, result.escenarios[1], result)

    if result.advertencias:
        _pdf_section(pdf, "Advertencias")
        for warning in result.advertencias:
            _pdf_multiline(pdf, f"- {warning}", align="L")

    _pdf_section(pdf, "Base de calculo")
    _pdf_multiline(
        pdf,
        "Haber individual = fondos de reserva + 5% del promedio anual de los "
        "ultimos 5 anos x tiempo de servicio. Pension mensual = haber ajustado / "
        "coeficiente C1 / 12. Fondo global = C2 x [(pension mensual x 12) + "
        "decimotercera + decimocuarta], validando el minimo legal del 50% de la "
        "remuneracion sectorial por anos de servicio.",
        align="L",
    )

    return pdf.output(dest="S").encode("latin-1")


def _write_scenario_summary(ws, start_row: int, scenario: ScenarioResult) -> None:
    rows = [
        ("Nombre", scenario.nombre),
        ("Pension mensual calculada", money(scenario.pension_mens_calculada)),
        ("Pension mensual aplicada", money(scenario.pension_mensual)),
        ("Limite aplicado", scenario.limite_aplicado),
        ("Decimotercera", money(scenario.decimotercera)),
        ("Decimocuarta", money(scenario.decimocuarta)),
        ("Fondo global calculado", money(scenario.fondo_global_calculado)),
        ("Minimo fondo global", money(scenario.minimo_fondo_global)),
        ("Fondo global aplicado", money(scenario.fondo_global)),
    ]
    for offset, (label, value) in enumerate(rows):
        ws.cell(start_row + offset, 1, label)
        ws.cell(start_row + offset, 2, value)


def _write_scenarios_sheet(ws, result: CalculationResult) -> None:
    headers = [
        "Escenario",
        "Descuento total",
        "Haber individual",
        "Haber ajustado",
        "C1",
        "Pension calculada",
        "Pension aplicada",
        "Limite",
        "C2",
        "Decimotercera",
        "Decimocuarta",
        "Fondo global calculado",
        "Minimo fondo global",
        "Fondo global aplicado",
    ]
    ws.append(headers)
    _format_header(ws, 1, len(headers))
    for scenario in result.escenarios:
        ws.append(
            [
                scenario.nombre,
                float(scenario.descuento_total),
                float(scenario.haber_individual),
                float(scenario.haber_ajustado),
                float(scenario.coeficiente_c1),
                float(scenario.pension_mens_calculada),
                float(scenario.pension_mensual),
                scenario.limite_aplicado,
                float(scenario.coeficiente_global.coeficiente),
                float(scenario.decimotercera),
                float(scenario.decimocuarta),
                float(scenario.fondo_global_calculado),
                float(scenario.minimo_fondo_global),
                float(scenario.fondo_global),
            ]
        )


def _write_inputs_sheet(ws, result: CalculationResult) -> None:
    ws.append(["Campo", "Valor"])
    _format_header(ws, 1, 2)
    rows = [
        ("Fondos reserva derecho", result.entrada.fondos_reserva_derecho),
        ("Fondos reserva pagados", result.entrada.fondos_reserva_pagados),
        ("Aportes patronales pagados", result.entrada.aportes_patronales_pagados),
        (
            "Decimotercera",
            result.entrada.decimotercera_remuneracion
            if result.entrada.decimotercera_remuneracion is not None
            else "Automatica: una pension mensual aplicada",
        ),
        ("Decimocuarta", result.entrada.decimocuarta_remuneracion),
        ("Remuneracion sectorial", result.entrada.remuneracion_sectorial),
        ("Doble jubilacion", "Si" if result.entrada.doble_jubilacion else "No"),
        ("Despido intempestivo", "Si" if result.entrada.despido_intempestivo else "No"),
    ]
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(["Remuneraciones ultimos 5 anos", "Valor"])
    _format_header(ws, ws.max_row, 2)
    for idx, value in enumerate(result.entrada.remuneraciones_ultimos_5, start=1):
        ws.append([f"Anio {idx}", float(value)])


def _write_sources_sheet(ws, result: CalculationResult) -> None:
    ws.append(["Fuente", "Detalle"])
    _format_header(ws, 1, 2)
    ws.append(["C1", result.coeficiente_c1_fuente])
    for scenario in result.escenarios:
        ws.append([f"C2 {scenario.nombre}", scenario.coeficiente_global.fuente])
    ws.append(["Resolucion 07-2021", "Limite maximo: promedio mensual del ultimo anio."])
    ws.append(["MDT-2016-0099", "Formula mensual y formula referencial de fondo global."])
    ws.append([
        "Resolucion CNJ 16-2025",
        "Precedente obligatorio: el escenario recomendado descuenta solo fondos de reserva pagados/depositados; aportes patronales quedan como comparativo.",
    ])
    ws.append([
        "Resolucion CNJ 04-2026",
        "Fondo global: aplicar acuerdos ministeriales vigentes a la fecha de terminacion laboral.",
    ])
    for title, detail in LEGAL_BASIS[:6]:
        ws.append([title, detail])


def _write_formulas_sheet(ws) -> None:
    ws["A1"] = "Formulas oficiales MDT-2016-0099"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A3"] = "Pension mensual"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "A = fondos de reserva; B = promedio anual ultimos 5 anos; C = tiempo de servicio; D = descuento; E = coeficiente C1."
    ws["A12"] = "Fondo global"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "A = coeficiente global C2; B = pension mensual; C = decimotercera; D = decimocuarta."
    if FORMULA_PENSION_PATH.exists():
        image = ExcelImage(str(FORMULA_PENSION_PATH))
        image.width = 297
        image.height = 114
        ws.add_image(image, "A5")
    if FORMULA_GLOBAL_PATH.exists():
        image = ExcelImage(str(FORMULA_GLOBAL_PATH))
        image.width = 249
        image.height = 103
        ws.add_image(image, "A14")


def _setup_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for idx in range(1, 12):
        ws.column_dimensions[get_column_letter(idx)].width = 22
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _format_header(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _pdf_section(pdf: FPDF, title: str) -> None:
    _ensure_space(pdf, 12)
    pdf.ln(4)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, _latin(title), ln=True)
    pdf.set_font("Arial", "", 10)


def _pdf_formula_section(pdf: FPDF) -> None:
    _pdf_section(pdf, "Formulas oficiales MDT-2016-0099")
    _ensure_space(pdf, 58)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 5, _latin("Pension mensual"), ln=True)
    if FORMULA_PENSION_PATH.exists():
        x = pdf.l_margin
        pdf.image(str(FORMULA_PENSION_PATH), x=x, y=pdf.get_y() + 1, w=75)
        pdf.set_y(pdf.get_y() + 31)
    pdf.set_font("Arial", "", 8.5)
    _pdf_multiline(
        pdf,
        "A = fondos de reserva; B = promedio anual ultimos 5 anos; C = tiempo de "
        "servicio; D = descuento; E = coeficiente C1.",
        align="L",
    )
    _ensure_space(pdf, 50)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(0, 5, _latin("Fondo global"), ln=True)
    if FORMULA_GLOBAL_PATH.exists():
        x = pdf.l_margin
        pdf.image(str(FORMULA_GLOBAL_PATH), x=x, y=pdf.get_y() + 1, w=65)
        pdf.set_y(pdf.get_y() + 29)
    pdf.set_font("Arial", "", 8.5)
    _pdf_multiline(
        pdf,
        "A = coeficiente global C2; B = pension mensual; C = decimotercera; "
        "D = decimocuarta.",
        align="L",
    )


def _pdf_pair(pdf: FPDF, label: str, value: str) -> None:
    _ensure_space(pdf, 7)
    label_width = 80
    row_height = 5.5
    start_x = pdf.l_margin
    start_y = pdf.get_y()
    value_x = start_x + label_width + 3
    value_width = pdf.w - pdf.r_margin - value_x

    pdf.set_xy(start_x, start_y)
    pdf.set_font("Arial", "B", 8.5)
    pdf.cell(label_width, row_height, _latin(label + ":"), border=0)

    pdf.set_xy(value_x, start_y)
    pdf.set_font("Arial", "", 8.5)
    pdf.multi_cell(value_width, row_height, _latin(value))
    if pdf.get_y() < start_y + row_height:
        pdf.set_y(start_y + row_height)


def _pdf_scenario(pdf: FPDF, scenario: ScenarioResult, result: CalculationResult) -> None:
    _pdf_pair(pdf, "Escenario", scenario.nombre)
    _pdf_pair(pdf, "Criterio", scenario.descripcion)
    _pdf_highlights(pdf, scenario)
    for label, value in [
        ("A - Fondos de reserva causados", _usd(result.entrada.fondos_reserva_derecho)),
        ("B - Promedio anual ultimos 5 anos", _usd(result.promedio_anual_ultimos_5)),
        ("C - Tiempo de servicio", f"{result.tiempo_servicio} anos"),
        ("D - Descuento aplicado", _usd(scenario.descuento_total)),
        ("E - Coeficiente C1", str(scenario.coeficiente_c1)),
        ("Limite mensual aplicado", scenario.limite_aplicado),
        ("Coeficiente global C2", str(scenario.coeficiente_global.coeficiente)),
        ("Decimotercera para fondo global", _usd(scenario.decimotercera)),
        ("Decimocuarta para fondo global", _usd(scenario.decimocuarta)),
        ("Minimo fondo global", _usd(scenario.minimo_fondo_global)),
    ]:
        _pdf_pair(pdf, label, value)
    pdf.ln(1)
    _pdf_multiline(
        pdf,
        "Formula pension mensual: ((A + (5% x B x C) - D) / E) / 12. "
        f"Sustitucion: (({result.entrada.fondos_reserva_derecho} + "
        f"(0.05 x {result.promedio_anual_ultimos_5} x {result.tiempo_servicio}) - "
        f"{scenario.descuento_total}) / {scenario.coeficiente_c1}) / 12 = "
        f"{money(scenario.pension_mens_calculada)} antes de limites.",
        align="L",
    )
    _pdf_multiline(
        pdf,
        "Formula fondo global: C2 x [(pension mensual aplicada x 12) + "
        "decimotercera + decimocuarta]. "
        f"Sustitucion: {scenario.coeficiente_global.coeficiente} x "
        f"[({money(scenario.pension_mensual)} x 12) + "
        f"{money(scenario.decimotercera)} + {money(scenario.decimocuarta)}] = "
        f"{money(scenario.fondo_global_calculado)}.",
        align="L",
    )


def _pdf_multiline(pdf: FPDF, text: str, align: str = "J") -> None:
    _ensure_space(pdf, 10)
    normalized = " ".join(str(text).split())
    if not normalized:
        pdf.ln(5)
        return
    _pdf_multicell(pdf, 0, 5, normalized, align=align)


def _pdf_multicell(pdf: FPDF, width: float, height: float, text: str, align: str = "L") -> None:
    pdf.multi_cell(width, height, _latin(text), align=align)


def _usd(value: Decimal) -> str:
    return f"USD {money(value):,.2f}"


def _latin(text: str) -> str:
    return (
        str(text)
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .encode("latin-1", errors="replace")
        .decode("latin-1")
    )


def _pdf_header(pdf: FPDF) -> None:
    start_y = 10
    title_x = 36
    if LOGO_PATH.exists():
        try:
            pdf.image(str(LOGO_PATH), x=10, y=start_y, w=18)
        except Exception:
            title_x = 10
    else:
        title_x = 10

    pdf.set_xy(title_x, start_y)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 7, _latin("Informe de calculo de Jubilacion Patronal"), ln=True)
    pdf.set_x(title_x)
    pdf.set_font("Arial", "", 9)
    pdf.multi_cell(
        pdf.w - pdf.r_margin - title_x,
        5,
        _latin(
            "Calculo referencial basado en el art. 216 del Codigo del Trabajo, "
            "MDT-2016-0099, Resolucion 07-2021 y tablas de coeficientes disponibles."
        ),
    )
    pdf.set_y(33)
    pdf.set_draw_color(31, 78, 121)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(2)


def _pdf_highlights(pdf: FPDF, scenario: ScenarioResult) -> None:
    _ensure_space(pdf, 46)
    cards = [
        ("Pension calculada", _usd(scenario.pension_mens_calculada)),
        ("Pension aplicada", _usd(scenario.pension_mensual)),
        ("Fondo global calculado", _usd(scenario.fondo_global_calculado)),
        ("Fondo global aplicado", _usd(scenario.fondo_global)),
    ]
    card_w = 88
    card_h = 17
    gap = 6
    x0 = pdf.l_margin
    y0 = pdf.get_y() + 1
    pdf.set_draw_color(31, 78, 121)

    for idx, (label, value) in enumerate(cards):
        col = idx % 2
        row = idx // 2
        x = x0 + col * (card_w + gap)
        y = y0 + row * (card_h + 3)
        pdf.set_fill_color(217, 234, 247)
        pdf.rect(x, y, card_w, card_h, style="DF")
        pdf.set_xy(x + 3, y + 2)
        pdf.set_text_color(31, 78, 121)
        pdf.set_font("Arial", "B", 8.5)
        pdf.cell(card_w - 6, 5, _latin(label), ln=True)
        pdf.set_xy(x + 3, y + 8)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(card_w - 6, 7, _latin(value))

    pdf.set_text_color(0, 0, 0)
    pdf.set_y(y0 + 2 * card_h + 7)


def _ensure_space(pdf: FPDF, needed_height: float) -> None:
    bottom = pdf.h - pdf.b_margin
    if pdf.get_y() + needed_height > bottom:
        pdf.add_page()
        _pdf_header(pdf)
